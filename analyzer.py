"""
analyzer.py — главный скрипт анализа крипто-гэмблинг сайтов.

Стратегия:
  1. Прямой fetch /terms страниц сайта → regex-экстракторы
  2. AskGamblers → рейтинги, жалобы, KYC время
  3. SSL + WHOIS → технические атрибуты
  4. Запись результатов в Excel

Запуск:
  python3 analyzer.py              # все сайты
  python3 analyzer.py --limit 10   # первые N сайтов
  python3 analyzer.py --start 20   # с позиции N
"""

import argparse
import re
import ssl
import time
import socket
from datetime import date, datetime, timezone

import httpx
import whois
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from scraper import scrape_site
from extractors import extract_all

# ---------------------------------------------------------------------------
# Конфигурация
# ---------------------------------------------------------------------------

INPUT_FILE = "Gembling_zadanie_Istomin_N.xlsx"
OUTPUT_FILE = "results/output.xlsx"
SHEET_NAME = "data"

# Маппинг: ключ из extract_all → колонка Excel (1-based)
COL_MAP = {
    "platform_name":       2,   # B
    "status_code":         3,   # C
    "platform_type":       4,   # D
    "is_AML":              5,   # E
    "is_KYC":              6,   # F
    "KYC_type":            7,   # G
    # "languages":         8,   # H — не извлекаем пока
    "web_archive_url":     9,   # I
    "legal_entity_name":  14,   # N
    "company_reg_number": 15,   # O
    "company_reg_country":16,   # P
    "license":            17,   # Q
    # Новые атрибуты — расширенные колонки начиная с R (18)
    "safety_score":       18,   # R
    "player_rating":      19,   # S
    "complaints_total":   20,   # T
    "complaints_resolved":21,   # U
    "payout_speed":       22,   # V
    "games_count":        23,   # W
    "supported_crypto":   24,   # X
    "supported_fiat":     25,   # Y
    "crypto_only":        26,   # Z
    "blockchains":        27,   # AA
    "is_decentralized":   28,   # AB
    "ssl_valid":          29,   # AC
    "domain_age_years":   30,   # AD
    "founded_year":       31,   # AE
}

TERMS_PATHS = [
    "/terms",
    "/terms-of-service",
    "/terms-and-conditions",
    "/legal",
    "/en/terms",
]

FETCH_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}


# ---------------------------------------------------------------------------
# Вспомогательные функции
# ---------------------------------------------------------------------------

def normalize_url(url: str) -> str:
    url = url.strip().rstrip("/")
    if not url.startswith("http"):
        url = "https://" + url
    return url


def fetch_text(url: str, timeout: int = 7) -> str | None:
    """Загружает URL и возвращает чистый текст."""
    try:
        r = httpx.get(url, headers=FETCH_HEADERS, timeout=timeout, follow_redirects=True)
        if r.status_code >= 400:
            return None
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "noscript", "nav", "header", "footer"]):
            tag.decompose()
        return " ".join(soup.get_text().split())
    except Exception:
        return None


def fetch_terms(base_url: str) -> str | None:
    """Перебирает стандартные пути для terms-страницы, возвращает первый рабочий текст."""
    for path in TERMS_PATHS:
        text = fetch_text(base_url + path)
        if text and len(text) > 200:
            return text
    return None


def fetch_askgamblers(site_name: str) -> dict:
    """Получает данные с AskGamblers: рейтинг надёжности, жалобы, скорость выплат и т.д."""
    slug = site_name.lower().replace(" ", "-").replace(".", "-")
    url = f"https://askgamblers.com/online-casinos/{slug}-casino-review"
    text = fetch_text(url)
    if not text:
        return {}

    result = {}

    # Safety / AskGamblers рейтинг
    for pattern in [
        r"AskGamblers\s+Rating[:\s]+(\d+(?:\.\d+)?)\s*/\s*10",
        r"Safety\s+Index[:\s]+(\d+(?:\.\d+)?)\s*/\s*10",
        r"Rating[:\s]+(\d+(?:\.\d+)?)\s*/\s*10",
    ]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            result["safety_score"] = m.group(1)
            break

    # Рейтинг игроков
    m = re.search(r"Player\s+Rating[:\s]+(\d+(?:\.\d+)?)", text, re.IGNORECASE)
    if m:
        result["player_rating"] = m.group(1)

    # Общее кол-во жалоб
    m = re.search(r"(\d+)\s+(?:total\s+)?complaint", text, re.IGNORECASE)
    if m:
        result["complaints_total"] = m.group(1)

    # Решённые жалобы
    m = re.search(r"[Rr]esolved[:\s]+(\d+)", text)
    if not m:
        m = re.search(r"(\d+)\s+(?:complaint.{0,20})?resolved", text, re.IGNORECASE)
    if m:
        result["complaints_resolved"] = m.group(1)

    # Скорость выплат
    for pattern in [
        r"[Pp]ayout\s+[Ss]peed[:\s]+([0-9\-]+\s*(?:hours?|days?))",
        r"[Ww]ithdrawal\s+[Tt]ime[:\s]+([0-9\-]+\s*(?:hours?|days?))",
        r"([0-9\-]+\s*(?:hours?|days?))\s+(?:payout|withdrawal)",
    ]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            result["payout_speed"] = m.group(1).strip()
            break

    # Кол-во игр
    for pattern in [
        r"([0-9][0-9,]+)\s*\+?\s*(?:casino\s+)?games?\b",
        r"[Gg]ames?\s+[Cc]ount[:\s]+([0-9,]+)",
        r"over\s+([0-9,]+)\s+games?\b",
    ]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            result["games_count"] = m.group(1).replace(",", "")
            break

    # Год основания
    m = re.search(r"[Ff]ounded\s+(?:in\s+)?[:\s]*(\d{4})", text)
    if m:
        result["founded_year"] = m.group(1)

    return result


def check_ssl(domain: str) -> str:
    """Проверяет валидность SSL-сертификата. Возвращает 'valid', 'invalid' или 'unknown'."""
    try:
        ctx = ssl.create_default_context()
        with ctx.wrap_socket(socket.socket(), server_hostname=domain) as s:
            s.settimeout(5)
            s.connect((domain, 443))
            cert = s.getpeercert()
            not_after = datetime.strptime(cert["notAfter"], "%b %d %H:%M:%S %Y %Z")
            return "valid" if not_after > datetime.now(timezone.utc).replace(tzinfo=None) else "expired"
    except ssl.SSLCertVerificationError:
        return "invalid"
    except Exception:
        return "unknown"


def get_domain_age(domain: str) -> str | None:
    """Возвращает возраст домена в годах через WHOIS."""
    try:
        w = whois.whois(domain)
        creation = w.creation_date
        if isinstance(creation, list):
            creation = creation[0]
        if isinstance(creation, datetime):
            age = (datetime.now(timezone.utc).replace(tzinfo=None) - creation.replace(tzinfo=None)).days / 365
            return f"{age:.1f}"
        if isinstance(creation, date):
            age = (date.today() - creation).days / 365
            return f"{age:.1f}"
    except Exception:
        pass
    return None


def guess_platform_name(url: str) -> str:
    """Извлекает название платформы из URL."""
    import re
    domain = re.sub(r"^https?://", "", url).split("/")[0]
    domain = re.sub(r"^www\.", "", domain)
    name = domain.split(".")[0]
    return name.capitalize()


# ---------------------------------------------------------------------------
# Основная логика обработки одного сайта
# ---------------------------------------------------------------------------

def analyze_site(url: str) -> dict:
    """Анализирует один сайт, возвращает dict с атрибутами."""
    url = normalize_url(url)
    result = {"url": url}

    # 1. Название платформы
    result["platform_name"] = guess_platform_name(url)

    # 2. Статус + homepage текст
    site_data = scrape_site(url)
    result["status_code"] = site_data["status"]
    result["web_archive_url"] = site_data["archive_url"]

    homepage_text = site_data["text"]

    # 3. Всегда пробуем terms страницу напрямую (независимо от статуса homepage)
    # Сайт может блокировать httpx на homepage, но отдавать /terms
    terms_text = fetch_terms(url)

    # 4. Если terms не получили — пробуем из архива
    if not terms_text and site_data["archive_url"]:
        archive_base = site_data["archive_url"].rstrip("/")
        for path in TERMS_PATHS[:3]:
            terms_text = fetch_text(archive_base + path)
            if terms_text and len(terms_text) > 200:
                break

    # 5. Если получили terms — обновляем статус (сайт скорее всего жив)
    if terms_text and site_data["status"] == "unreachable":
        site_data["status"] = "active"
        result["status_code"] = "active"

    # 5. Объединяем тексты для анализа
    combined_text = (terms_text or "") + "\n\n" + homepage_text

    # 6. Regex-экстракция
    attrs = extract_all(combined_text, site_data["status"], site_data["archive_url"])
    result.update(attrs)

    # 7. AskGamblers (рейтинги, жалобы, скорость выплат, кол-во игр)
    ag_data = fetch_askgamblers(result["platform_name"])
    result.update(ag_data)

    # 8. SSL валидность
    domain = re.sub(r"^https?://", "", url).split("/")[0]
    result["ssl_valid"] = check_ssl(domain)

    # 9. Возраст домена через WHOIS
    result["domain_age_years"] = get_domain_age(domain)

    return result


# ---------------------------------------------------------------------------
# Чтение / запись Excel
# ---------------------------------------------------------------------------

def load_urls() -> list[tuple[int, str]]:
    """Читает URL из листа data, возвращает список (row_number, url)."""
    wb = load_workbook(INPUT_FILE)
    ws = wb[SHEET_NAME]
    urls = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        url = row[0]
        if url and isinstance(url, str):
            urls.append((i, url.strip()))
    return urls


def ensure_headers(ws):
    """Добавляет заголовки для новых колонок если их ещё нет."""
    for key, col in COL_MAP.items():
        if ws.cell(row=1, column=col).value is None:
            ws.cell(row=1, column=col, value=key)


def save_result(ws, row: int, data: dict):
    """Записывает результаты в строку Excel."""
    for key, col in COL_MAP.items():
        value = data.get(key)
        if value is not None:
            ws.cell(row=row, column=col, value=str(value))


# ---------------------------------------------------------------------------
# Точка входа
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=None, help="Максимум сайтов")
    parser.add_argument("--start", type=int, default=0, help="Начать с позиции N")
    args = parser.parse_args()

    # Подготовка output файла — копируем только если не существует
    import shutil, os
    os.makedirs("results", exist_ok=True)
    if not os.path.exists(OUTPUT_FILE):
        shutil.copy(INPUT_FILE, OUTPUT_FILE)
        print(f"Создан новый файл результатов: {OUTPUT_FILE}")

    wb = load_workbook(OUTPUT_FILE)
    ws = wb[SHEET_NAME]
    ensure_headers(ws)

    urls = load_urls()
    urls = urls[args.start:]
    if args.limit:
        urls = urls[:args.limit]

    print(f"Обрабатываем {len(urls)} сайтов...\n")

    for i, (row_num, url) in enumerate(urls):
        print(f"[{i+1}/{len(urls)}] {url}")
        try:
            data = analyze_site(url)
            save_result(ws, row_num, data)
            wb.save(OUTPUT_FILE)

            print(
                f"  status={data.get('status_code')} | type={data.get('platform_type')} "
                f"| KYC={data.get('KYC_type')} | AML={data.get('is_AML')} "
                f"| country={data.get('company_reg_country')} "
                f"| ssl={data.get('ssl_valid')} | age={data.get('domain_age_years') or '?'}y "
                f"| crypto={data.get('supported_crypto', '-')} "
                f"| defi={data.get('is_decentralized', '-')}"
            )
        except Exception as e:
            print(f"  ERROR: {e}")

        # Пауза чтобы не спамить серверы
        time.sleep(1.5)

    print(f"\nГотово! Результаты: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
